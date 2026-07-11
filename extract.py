"""Flatten the messy 股票明細 ledger into the rev-3 LOT-BASED schema (design.md §A.0/§A.1).

Outputs (real data — distinct from test fixtures/ which stay synthetic):
  - data/held_lots.json  data/trades.json  data/deposits.json  data/dividends.json
  - portfolio_normalized.xlsx   (tabs: HeldLots, Trades, Deposits, Dividends, Prices)
  - data/*.csv                  (Google Sheets import)

Classification (authoritative, design.md §A.1) — projections are independent, not mutually exclusive
(a 尚未交易=Y buy row is BOTH a Trade[buy] and a HeldLot):
  Dividend ⇔ 股票 == "股息"                                  (code+name from 明細)
  Deposit  ⇔ 項目 == "CD轉入"                                 (amount = 存入)
  Trade buy  ⇔ 項目=="轉帳支取" & 股票 set & != "股息"          (amount = 支出)
  Trade sell ⇔ 項目=="轉帳存入" & 股票 set & != "股息"          (amount = 存入)
  HeldLot  ⇔ 尚未交易 starts "Y" & 股票 != "股息"  (shares=股數, buyPrice=股價, cost=目前投資金額)
"""
import json, os, re, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import warnings
warnings.filterwarnings("ignore")

SRC = "股票 2023.xlsx"
DIV = "股息"
os.makedirs("data", exist_ok=True)

KNOW_CODES = {
    "台達電": "2308",
    "國產": "2504",
    "康舒": "6282",
    "星宇航空": "2646",
    "晶心科": "6533",
    "晶電": "2448",
    "瑞昱": "2379",
    "統懋": "2434",
    "花王": "8906",
    "華冠": "8101",
    "華邦電": "2344",
    "錦明": "3230",
}


def load_ledger():
    df = pd.read_excel(SRC, sheet_name="股票明細", header=1).dropna(how="all")
    df = df[df["date"].notna()].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    return df[df["date"].notna()]


def parse_div(detail):
    s = str(detail) if pd.notna(detail) else ""
    m = re.search(r"(\d{4})", s)
    code = m.group(1) if m else None
    name = re.sub(r"現金股息|股票股利|股利|\d+", "", s).strip()
    return code, name


def num(v):
    return float(v) if pd.notna(v) else None


def extract(df):
    held, trades, deposits, divs, name2code = [], [], [], [], {}
    for _, r in df.iterrows():
        ticker, action = r["股票"], r["項目"]
        d = r["date"].strftime("%Y-%m-%d")
        is_div = ticker == DIV

        if is_div:
            code, name = parse_div(r["明細"])
            if code and name:
                name2code.setdefault(name, code)
            divs.append({"date": d, "code": code, "name": name,
                         "amount": float(r["存入"]), "detail": str(r["明細"])})
        if action == "CD轉入" and pd.notna(r["存入"]):
            deposits.append({"date": d, "amount": float(r["存入"])})
        if action == "轉帳支取" and pd.notna(ticker) and not is_div and pd.notna(r["支出"]):
            trades.append({"date": d, "type": "buy", "ticker": str(ticker),
                           "shares": num(r["股數"]), "amount": float(r["支出"])})
        if action == "轉帳存入" and pd.notna(ticker) and not is_div:
            amt = num(r["存入"])                       # corrupt rows (e.g. 2023-12-12 威剛) have blank 存入
            if amt is not None:
                trades.append({"date": d, "type": "sell", "ticker": str(ticker),
                               "shares": num(r["股數"]), "amount": amt})
        if str(r["尚未交易"]).upper().startswith("Y") and pd.notna(ticker) and not is_div:
            held.append({"date": d, "ticker": str(ticker), "shares": num(r["股數"]),
                         "buyPrice": num(r["股價"]), "cost": float(num(r["目前投資金額"]) or 0)})
    # Merge known codes
    for name, code in KNOW_CODES.items():
        name2code.setdefault(name, code)
    return held, trades, deposits, divs, name2code


def holdings(held):
    pos = {}
    for h in held:
        if h["shares"] is not None:
            pos[h["ticker"]] = pos.get(h["ticker"], 0) + h["shares"]
    return pos


def dump(obj, path):
    # allow_nan=False → fail loud rather than emit invalid JSON (NaN breaks JSON.parse)
    json.dump(obj, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=2, allow_nan=False)


def write_normalized(held, trades, deposits, divs, codes):
    wb = Workbook(); bold = Font(bold=True)
    tabs = {
        "HeldLots": (["date", "ticker", "shares", "buyPrice", "cost"], held),
        "Trades":   (["date", "type", "ticker", "shares", "amount"], trades),
        "Deposits": (["date", "amount"], deposits),
        "Dividends":(["date", "code", "name", "amount", "detail"], divs),
    }
    first = True
    for name, (cols, rows) in tabs.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first: ws.title = name; first = False
        ws.append(cols)
        for row in rows:
            ws.append([row[c] for c in cols])
        for c in ws[1]: c.font = bold
        pd.DataFrame(rows, columns=cols).to_csv(f"data/{name}.csv", index=False, encoding="utf-8-sig")

    # Extract all active years from the dataset to generate year-end columns
    all_dates = [x["date"] for x in held + trades + deposits + divs if "date" in x]
    years = sorted(list(set(int(d[:4]) for d in all_dates if len(d) >= 4 and d[:4].isdigit())))
    from datetime import date
    current_year = date.today().year
    years = [y for y in years if 2010 <= y <= current_year]
    year_cols = [f"{y}-12-31" for y in years]

    wp = wb.create_sheet("Prices")
    wp.append(["ticker", "code", "price", "closeyest"] + year_cols)

    for r_idx, tk in enumerate(sorted(holdings(held)), start=2):
        code = codes.get(tk)
        row_data = [tk, code or ""]
        if code:
            row_data.append(f'=IFERROR(GOOGLEFINANCE("TPE:{code}"), GET_TAIWAN_STOCK_PRICE("{code}"))')
            row_data.append(f'=IFERROR(GOOGLEFINANCE("TPE:"&$B{r_idx},"closeyest"), GET_TAIWAN_STOCK_PRICE($B{r_idx}, "closeyest"))')
            for y in years:
                row_data.append(
                    f'=IFERROR(INDEX(GOOGLEFINANCE("TPE:"&$B{r_idx},"close",DATE({y},12,31)),2,2), '
                    f'GET_TAIWAN_STOCK_PRICE($B{r_idx}, "{y}-12-31"))'
                )
        else:
            row_data.append("")
            row_data.append("")
            for y in years:
                row_data.append("")
        wp.append(row_data)

    for c in wp[1]: c.font = bold
    wb.save("portfolio_normalized.xlsx")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    df = load_ledger()
    held, trades, deposits, divs, codes = extract(df)

    dump(held, "data/held_lots.json"); dump(trades, "data/trades.json")
    dump(deposits, "data/deposits.json"); dump(divs, "data/dividends.json")

    # Stand-in prices for OFFLINE preview only (avg cost/share, keyed by today).
    # Live runtime replaces this with GOOGLEFINANCE via the Apps Script source.
    from datetime import date
    today = date.today().isoformat()
    agg = {}
    for h in held:
        if h["shares"]:
            a = agg.setdefault(h["ticker"], [0, 0]); a[0] += h["cost"]; a[1] += h["shares"]
    dump({tk: {today: round(c / s, 2)} for tk, (c, s) in agg.items() if s}, "data/prices.json")

    write_normalized(held, trades, deposits, divs, codes)

    # verification (design.md DX-2/3/4)
    assert abs(sum(d["amount"] for d in divs) - df[df["股票"] == DIV]["存入"].sum()) < 1e-6, "DX-2"
    assert all(t["ticker"] != DIV for t in trades) and all(h["ticker"] != DIV for h in held), "DX-3"

    pos = holdings(held)
    print(f"held lots: {len(held)} ({len(pos)} tickers)   trades: {len(trades)}   "
          f"deposits: {len(deposits)}   dividends: {len(divs)}")
    print(f"invested capital (Σ deposits): NT$ {sum(x['amount'] for x in deposits):,.0f}")
    print(f"cost of holdings (Σ HeldLot.cost): NT$ {sum(h['cost'] for h in held):,.0f}")
    print(f"dividends total: NT$ {sum(d['amount'] for d in divs):,.0f}  (DX-2 ✓, DX-3 ✓)")
    nullsh = [h["ticker"] for h in held if h["shares"] is None]
    print(f"null-share lots (review): {len(nullsh)} {sorted(set(nullsh))}")
    print(f"held tickers without TWSE code: {sorted(t for t in pos if t not in codes)}")
    print("Wrote data/*.json + portfolio_normalized.xlsx (HeldLots/Trades/Deposits/Dividends/Prices)")
